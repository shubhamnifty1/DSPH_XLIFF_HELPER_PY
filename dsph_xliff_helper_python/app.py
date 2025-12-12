import sys
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from flask import Flask, request, render_template, send_file
import os
import xmltodict
import openpyxl
import logging
from datetime import datetime

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
LOG_FOLDER = "logs"
LOG_FILE_PATH = os.path.join(LOG_FOLDER, "app.log")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(LOG_FOLDER, exist_ok=True)


# -----------------------------------------
# CLEAR LOG FILE BEFORE ANYTHING WRITES
# -----------------------------------------
open("logs/app.log", "w", encoding="utf-8").close()



# -----------------------------
# LOGGING SETUP (UTF-8 ENABLED)
# -----------------------------
logging.basicConfig(
    handlers=[logging.FileHandler(LOG_FILE_PATH, encoding="utf-8")],
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

console = logging.StreamHandler()
console.setLevel(logging.INFO)
console.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
logging.getLogger().addHandler(console)


def log_separator():
    """Add visual separator + keep only last 3 runs."""
    with open(LOG_FILE_PATH, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Keep last 1000 lines temporarily
    lines = lines[-1000:]

    # Count runs
    run_starts = [i for i, line in enumerate(lines) if "NEW RUN" in line]

    if len(run_starts) >= 3:
        # Trim everything before the 3rd last run
        cutoff = run_starts[-3]
        lines = lines[cutoff:]

    with open(LOG_FILE_PATH, "w", encoding="utf-8") as f:
        f.writelines(lines)

    # Write separator
    with open(LOG_FILE_PATH, "a", encoding="utf-8") as f:
        f.write("\n----------------------------------------\n")
        f.write(f"🕒 NEW RUN: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("----------------------------------------\n\n")


# -----------------------------------------
# HOME PAGE
# -----------------------------------------
@app.route("/")
def index():
    logging.info("📄 Home page opened")
    return render_template("index.html")


# -----------------------------------------
# STEP 1 — XLF → EXCEL
# -----------------------------------------
@app.route("/convert_xlf_to_excel", methods=["POST"])
def convert_xlf_to_excel():

    log_separator()
    logging.info("📁 STEP 1: XLF → Excel started")

    file = request.files.get("xlf_file")
    if not file:
        logging.error("❌ No XLF file uploaded")
        return "No file uploaded", 400

    original_filename = file.filename
    xlf_path = os.path.join(UPLOAD_FOLDER, "source.xlf")
    file.save(xlf_path)

    logging.info(f"📥 XLF uploaded: {original_filename}")

    base_name = os.path.splitext(original_filename)[0]  # ex: AM_REPORT_DE
    parts = base_name.split("_")

    model_name = "_".join(parts[:-1]) if len(parts) >= 2 else "MODEL"
    lang_code = parts[-1].upper() if len(parts) >= 2 and len(parts[-1]) == 2 else ""

    logging.info(f"🆔 Model name: {model_name}")
    logging.info(f"🌐 Language code: {lang_code}")

    with open(xlf_path, "r", encoding="utf-8") as f:
        xlf_dict = xmltodict.parse(f.read())

    trans_units = xlf_dict["xliff"]["file"]["body"]["trans-unit"]

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append(["id", "source", "target", "lang", "auto_translate"])

    row = 2

    for unit in trans_units:
        full_id = unit["@id"]
        src = unit.get("source", "")
        if isinstance(src, list):
            src = src[0]

        if "#" in full_id and "@" in full_id:
            field_id = full_id.split("#")[1].split("@")[0]
        else:
            field_id = full_id

        formula = f'=IF(B{row}="", "", TRANSLATE(B{row}, "en", D{row}))'

        sheet.append([field_id, src, "", lang_code, formula])
        row += 1

    excel_path = os.path.join(OUTPUT_FOLDER, f"{model_name}.xlsx")
    workbook.save(excel_path)

    logging.info(f"📤 Excel created: {excel_path}")

    return send_file(excel_path, as_attachment=True, download_name=f"{model_name}.xlsx")


# -----------------------------------------
# STEP 2 — EXCEL → XLF
# -----------------------------------------
@app.route("/convert_excel_to_xlf", methods=["POST"])
def convert_excel_to_xlf():

    log_separator()
    logging.info("📁 STEP 2: Excel → XLF started")

    file = request.files.get("excel_file")
    if not file:
        logging.error("❌ No Excel file uploaded")
        return "No file uploaded", 400

    excel_path = os.path.join(UPLOAD_FOLDER, "input.xlsx")
    file.save(excel_path)
    logging.info("📥 Excel uploaded")

    # Load model name from Excel filename
    excel_filename = file.filename  # ex: AM_REPORT.xlsx
    base_model_name = os.path.splitext(excel_filename)[0]

    logging.info(f"🆔 Model from Excel file: {base_model_name}")

    original_xlf_path = os.path.join(UPLOAD_FOLDER, "source.xlf")
    if not os.path.exists(original_xlf_path):
        logging.error("⚠️ source.xlf missing — Step 1 was not executed before Step 2.")
        return "Run XLF → Excel first.", 400

    with open(original_xlf_path, "r", encoding="utf-8") as f:
        original_dict = xmltodict.parse(f.read())

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    translation_map = {}
    lang_code = ""

    for row in sheet.iter_rows(min_row=2, values_only=True):
        field_id, src, tgt, lang_cell, auto = row

        if not field_id:
            continue

        if lang_cell and len(str(lang_cell)) == 2:
            lang_code = str(lang_cell).upper()

        translation_map[field_id] = tgt if (tgt and str(tgt).strip()) else src

    logging.info(f"🔄 Loaded {len(translation_map)} translations")

    trans_units = original_dict["xliff"]["file"]["body"]["trans-unit"]

    for unit in trans_units:
        full_id = unit["@id"]

        if "#" in full_id and "@" in full_id:
            field_id = full_id.split("#")[1].split("@")[0]
        else:
            field_id = full_id

        unit["target"] = translation_map.get(field_id, unit.get("source", ""))

    output_filename = f"{base_model_name}_{lang_code}.xlf"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(xmltodict.unparse(original_dict, pretty=True))

    logging.info(f"📤 Final XLF created: {output_path}")

    # --------------------
    # CLEANUP
    # --------------------
    logging.info("🧹 Cleanup...")

    for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
        for fname in os.listdir(folder):
            fpath = os.path.join(folder, fname)

            if fpath == output_path:
                continue

            try:
                os.remove(fpath)
                logging.info(f"🗑️ Deleted {fpath}")
            except Exception as e:
                logging.error(f"❌ Could not delete {fpath}: {e}")

    return send_file(output_path, as_attachment=True, download_name=output_filename)


# -----------------------------------------
# RUN SERVER
# -----------------------------------------
if __name__ == "__main__":
    logging.info("🌍 Server running at http://localhost:5000")
    app.run(debug=True)
